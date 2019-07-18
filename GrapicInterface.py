from openpyxl import *
from tkinter import *
  
# globally declare wb and sheet variable 
  
# opening the existing excel file 
wb = load_workbook('/home/seethalprince/Desktop/cdc1.xlsx') 
  
# create the sheet object 
sheet = wb.active 
  
def excel(): 
      
    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['f'].width = 10
    
    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Age"
    sheet.cell(row=1, column=3).value = "BodyTemp"
    sheet.cell(row=1, column=4).value = "Pressure"
    sheet.cell(row=1, column=5).value = "BloodLevel"
    #sheet.cell(row=1, column=6).value = "BodyTemp"
    #sheet.cell(row=1, column=7).value = "Pressure"
    
    
    
  
# Function to set focus (cursor) 
def focus1(event): 
    # set focus on the course_field box 
    name_field.focus_set() 
  
  
# Function to set focus 
def focus2(event): 
    # set focus on the sem_field box 
    age_field.focus_set() 
  
  
# Function to set focus 
def focus3(event): 
    # set focus on the form_no_field box 
    bodytemp_field.focus_set() 
  
  
# Function to set focus 
def focus4(event): 
    # set focus on the contact_no_field box 
    pressure_field.focus_set() 
def focus5(event): 
    # set focus on the contact_no_field box 
    blood_field.focus_set() 
    
# Function to set focus 

  
# Function to set focus 

  
# Function for clearing the 
# contents of text entry boxes 
def clear(): 
      
    # clear the content of text entry box 
    name_field.delete(0, END) 
    age_field.delete(0, END) 
    bodytemp_field.delete(0, END) 
    pressure_field.delete(0, END) 
    blood_field.delete(0, END)
  
# Function to take data from GUI  
# window and write to an excel file 
def insert(): 
      
    # if user not fill any entry 
    # then print "empty input" 
    if (name_field.get() == "" and
        age_field.get() == "" and
        bodytemp_field.get() == "" and
        pressure_field.get() == "" and
    blood_field.get()==""): 
              
        print("empty input") 
  
    else: 
  
        # assigning the max row and max column 
        # value upto which data is written 
        # in an excel sheet to the variable 
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
        # get method returns current text 
        # as string which we write into 
        # excel spreadsheet at particular location 
        sheet.cell(row=current_row + 1, column=1).value = name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = age_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = bodytemp_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = pressure_field.get() 
        sheet.cell(row=current_row + 1, column=4).value = blood_field.get() 
       
        # save the file 
        #wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx') 
    wb.save('/home/seethalprince/Desktop/cdc1.xlsx') 
        # set focus on the name_field box 
    name_field.focus_set() 
  
        # call the clear() function 
    clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk() 
  
    # set the background colour of GUI window 
    root.configure(background='light green') 
  
    # set the title of GUI window 
    root.title("registration form") 
  
    # set the configuration of GUI window 
    root.geometry("500x300") 
  
    excel() 
  
    # create a Form label 
    heading = Label(root, text="Form", bg="light green") 
  
    # create a Name label 
    name = Label(root, text="Name", bg="light green") 
  
    # create a Course label 
    age = Label(root, text="Age", bg="light green") 
  
    # create a Semester label 
    bodytemp= Label(root, text="BodyTemp", bg="light green") 
  
    # create a Form No. lable 
    pressure= Label(root, text="Pressure", bg="light green") 
  
    # create a Contact No. label 
    blood = Label(root, text="Blod Level.", bg="light green") 
  
   
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    age.grid(row=2, column=0) 
    bodytemp.grid(row=3, column=0) 
    pressure.grid(row=4, column=0) 
    blood.grid(row=5,column=0)
   
  
    # create a text entry box 
    # for typing the information 
    name_field = Entry(root) 
    age_field = Entry(root) 
    bodytemp_field = Entry(root) 
    pressure_field = Entry(root) 
    blood_field = Entry(root)
    
    # bind method of widget is used for 
    # the binding the function with the events 
  
    # whenever the enter key is pressed 
    # then call the focus1 function 
    name_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    age_field.bind("<Return>", focus2) 
  
    # whenever the enter key is pressed 
    # then call the focus3 function 
    bodytemp_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus4 function 
    pressure_field.bind("<Return>", focus4) 
    blood_field.bind( "<Return>",focus5)
  
    #
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    name_field.grid(row=1, column=1, ipadx="100") 
    age_field.grid(row=2, column=1, ipadx="100") 
    bodytemp_field.grid(row=3, column=1, ipadx="100") 
    pressure_field.grid(row=4, column=1, ipadx="100") 
    blood_field.grid(row=5,column=1,ipadx="100")
   
    # call excel function 
    excel() 
  
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=8, column=1) 
  
    # start the GUI 
    root.mainloop() 

